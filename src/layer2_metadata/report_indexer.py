"""과거 사고보고서 임베딩 및 유사 사례 검색."""
from __future__ import annotations
import sys as _sys; from pathlib import Path as _Path
_sys.path.insert(0, str(_Path(__file__).resolve().parent.parent))

import hashlib
import json
import logging
from pathlib import Path
from typing import Any

import duckdb
import numpy as np
from openpyxl import load_workbook

from config_new import DUCKDB_PATH
from .db_schema import init_db

logger = logging.getLogger(__name__)

# ------------------------------------------------------------------
# 임베딩 백엔드 선택
# ------------------------------------------------------------------

_USE_SBERT: bool = False
_sbert_model: Any = None

try:
    from sentence_transformers import SentenceTransformer

    _USE_SBERT = True
except ImportError:
    logger.info("sentence-transformers 미설치 -- TF-IDF fallback 사용")


class ReportIndexer:
    """report_archive 테이블에 보고서를 적재하고 유사 사례를 검색한다."""

    def __init__(self, db_path: str = DUCKDB_PATH) -> None:
        self._conn: duckdb.DuckDBPyConnection = init_db(db_path)
        self._sbert: Any = None

        if _USE_SBERT:
            global _sbert_model
            if _sbert_model is None:
                _sbert_model = SentenceTransformer("all-MiniLM-L6-v2")
            self._sbert = _sbert_model

        # TF-IDF fallback 은 search 시 lazy 초기화
        self._tfidf_vectorizer: Any = None
        self._tfidf_matrix: Any = None
        self._tfidf_ids: list[str] = []

    # ------------------------------------------------------------------
    # 엑셀 적재
    # ------------------------------------------------------------------
    def index_xlsx(self, xlsx_path: str | Path) -> int:
        """엑셀 보고서를 읽어 report_archive 에 적재하고 임베딩을 생성한다.

        Parameters
        ----------
        xlsx_path : str | Path
            표준화데이터 시트를 포함한 엑셀 경로.

        Returns
        -------
        int
            적재된 행 수.
        """
        xlsx_path = Path(xlsx_path)
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)

        # '표준화데이터' 시트 우선, 없으면 첫 번째 시트
        if "표준화데이터" in wb.sheetnames:
            ws = wb["표준화데이터"]
        else:
            ws = wb[wb.sheetnames[0]]

        rows_iter = ws.iter_rows(values_only=True)
        header: tuple[Any, ...] = next(rows_iter)
        col_map: dict[str, int] = {
            str(h).strip(): i for i, h in enumerate(header) if h is not None
        }

        count = 0
        for row in rows_iter:
            if row is None or all(c is None for c in row):
                continue
            record = self._parse_row(col_map, row)
            if record is None:
                continue

            text_repr = self._record_to_text(record)
            embedding = self._generate_embedding(text_repr)

            self._upsert_archive(record, embedding)
            count += 1

        wb.close()

        # TF-IDF fallback 캐시 무효화
        self._tfidf_matrix = None

        logger.info("index_xlsx: %d rows loaded from %s", count, xlsx_path.name)
        return count

    # ------------------------------------------------------------------
    # 유사 사례 검색
    # ------------------------------------------------------------------
    def search_similar(self, query_text: str | None = None, top_k: int = 3,
                       accident_type: str | None = None,
                       road_name: str | None = None) -> list[dict]:
        """코사인 유사도로 유사 사례 상위 top_k 건을 반환한다.

        Parameters
        ----------
        query_text : str | None
            검색 쿼리 (예: "경부 부산 추돌 화물차 안전거리"). None이면
            accident_type/road_name으로 조립(report_generator 호환 — 기존 시그니처
            불일치로 similar_cases가 항상 빈 리스트였던 버그 수정).
        top_k : int
            반환 건수.

        Returns
        -------
        list[dict]
            report_id, score, road_name, direction, accident_type,
            cause, vehicles_summary 등을 담은 dict 목록.
        """
        if query_text is None:
            query_text = " ".join(str(x) for x in (road_name, accident_type) if x)
        if not query_text:
            return []
        if self._sbert is not None:
            return self._search_sbert(query_text, top_k)
        return self._search_tfidf(query_text, top_k)

    # ------------------------------------------------------------------
    # SBERT 검색
    # ------------------------------------------------------------------
    def _search_sbert(self, query_text: str, top_k: int) -> list[dict]:
        query_emb = self._sbert.encode(query_text, normalize_embeddings=True)

        rows = self._conn.execute(
            "SELECT report_id, road_name, direction, accident_type, "
            "cause, vehicles_summary, full_record FROM report_archive"
        ).fetchall()

        if not rows:
            return []

        # report_archive 에 embedding 컬럼이 있으면 사용, 없으면 재생성
        ids: list[str] = []
        embeddings: list[np.ndarray] = []
        meta: list[dict] = []

        for r in rows:
            rid = r[0]
            ids.append(rid)
            meta.append({
                "report_id": rid,
                "road_name": r[1],
                "direction": r[2],
                "accident_type": r[3],
                "cause": r[4],
                "vehicles_summary": r[5],
            })
            # 텍스트 재생성 -> 임베딩 (DB 에 float[] 저장이 어려우므로)
            text_repr = " ".join(str(v) for v in r[:6] if v)
            emb = self._sbert.encode(text_repr, normalize_embeddings=True)
            embeddings.append(emb)

        emb_matrix = np.stack(embeddings)  # (N, dim)
        scores = emb_matrix @ query_emb  # cosine similarity (normalized)

        top_idx = np.argsort(scores)[::-1][:top_k]
        results: list[dict] = []
        for idx in top_idx:
            entry = meta[idx].copy()
            entry["score"] = round(float(scores[idx]), 4)
            results.append(entry)
        return results

    # ------------------------------------------------------------------
    # TF-IDF fallback 검색
    # ------------------------------------------------------------------
    def _search_tfidf(self, query_text: str, top_k: int) -> list[dict]:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.metrics.pairwise import cosine_similarity

        rows = self._conn.execute(
            "SELECT report_id, road_name, direction, accident_type, "
            "cause, vehicles_summary FROM report_archive"
        ).fetchall()

        if not rows:
            return []

        corpus: list[str] = []
        meta: list[dict] = []
        for r in rows:
            text = " ".join(str(v) for v in r if v)
            corpus.append(text)
            meta.append({
                "report_id": r[0],
                "road_name": r[1],
                "direction": r[2],
                "accident_type": r[3],
                "cause": r[4],
                "vehicles_summary": r[5],
            })

        vectorizer = TfidfVectorizer()
        tfidf_matrix = vectorizer.fit_transform(corpus)
        query_vec = vectorizer.transform([query_text])
        scores = cosine_similarity(query_vec, tfidf_matrix).flatten()

        top_idx = np.argsort(scores)[::-1][:top_k]
        results: list[dict] = []
        for idx in top_idx:
            entry = meta[idx].copy()
            entry["score"] = round(float(scores[idx]), 4)
            results.append(entry)
        return results

    # ------------------------------------------------------------------
    # 임베딩 생성
    # ------------------------------------------------------------------
    def _generate_embedding(self, text: str) -> list[float]:
        """텍스트에서 임베딩 벡터를 생성한다.

        sentence-transformers 가 있으면 all-MiniLM-L6-v2 (384차원),
        없으면 빈 리스트를 반환한다 (DB 적재 시 null).
        """
        if self._sbert is not None:
            vec = self._sbert.encode(text, normalize_embeddings=True)
            return vec.tolist()
        return []

    # ------------------------------------------------------------------
    # 내부 유틸
    # ------------------------------------------------------------------
    def _parse_row(
        self, col_map: dict[str, int], row: tuple[Any, ...]
    ) -> dict[str, Any] | None:
        """엑셀 행 하나를 report_archive 컬럼에 매핑한다."""

        def _get(name: str) -> Any:
            idx = col_map.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        # 노선명이 없으면 유효하지 않은 행
        road_name = _get("노선명")
        if not road_name:
            return None

        # 피해차량 1~10 을 요약
        vehicles: list[str] = []
        for i in range(1, 11):
            v = _get(f"피해차량_{i}")
            if v:
                vehicles.append(str(v).strip())
        vehicles_summary = "+".join(vehicles) if vehicles else None

        casualties_total = _get("인명피해(총수)")
        fatalities = _get("사망자수")
        congestion_km = _get("정체길이")

        # report_id: 노선+날짜+이정 기반 해시
        date_val = _get("사고접보 시각(날짜)")
        raw_key = f"{road_name}_{date_val}_{_get('이정')}_{_get('사고원인')}"
        report_id = hashlib.md5(raw_key.encode()).hexdigest()[:16]

        # full_record: 전체 컬럼을 JSON 으로 저장
        full_record: dict[str, Any] = {}
        for col_name, idx in col_map.items():
            if idx < len(row):
                val = row[idx]
                if val is not None:
                    full_record[col_name] = str(val)

        return {
            "report_id": report_id,
            "source": "도공",
            "date": date_val,
            "road_name": str(road_name),
            "direction": _get("방향"),
            "km_post": _get("이정"),
            "point_type": _get("사고지점 유형"),
            "weather": _get("기상"),
            "accident_type": _get("사고 유형"),
            "cause": _get("사고원인"),
            "vehicles_summary": vehicles_summary,
            "casualties_total": _safe_int(casualties_total),
            "fatalities": _safe_int(fatalities),
            "congestion_km": _safe_int(congestion_km),
            "full_record": full_record,
        }

    @staticmethod
    def _record_to_text(record: dict[str, Any]) -> str:
        """레코드를 임베딩용 텍스트 문자열로 변환한다."""
        parts: list[str] = []
        for key in (
            "road_name", "direction", "accident_type", "cause",
            "vehicles_summary", "weather", "point_type",
        ):
            val = record.get(key)
            if val:
                parts.append(str(val))
        return " ".join(parts)

    def _upsert_archive(
        self, record: dict[str, Any], embedding: list[float]
    ) -> None:
        """report_archive 테이블에 INSERT OR REPLACE 한다."""
        full_json = json.dumps(
            record.get("full_record", {}), ensure_ascii=False
        )

        sql = """
            INSERT OR REPLACE INTO report_archive (
                report_id, source, date, road_name, direction,
                km_post, point_type, weather, accident_type, cause,
                vehicles_summary, casualties_total, fatalities,
                congestion_km, full_record
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        self._conn.execute(sql, [
            record["report_id"],
            record.get("source"),
            record.get("date"),
            record.get("road_name"),
            record.get("direction"),
            record.get("km_post"),
            record.get("point_type"),
            record.get("weather"),
            record.get("accident_type"),
            record.get("cause"),
            record.get("vehicles_summary"),
            record.get("casualties_total"),
            record.get("fatalities"),
            record.get("congestion_km"),
            full_json,
        ])


# ------------------------------------------------------------------
# 유틸
# ------------------------------------------------------------------

def _safe_int(val: Any) -> int | None:
    """안전하게 정수 변환. 실패 시 None."""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None
